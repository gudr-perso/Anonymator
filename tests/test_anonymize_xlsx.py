from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from anonymator.referential import Referential
from anonymator.ner import FakeNer
from anonymator.files.anonymize_file import anonymize_xlsx

def _make_book(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance"
    ws["A1"] = "Libellé"; ws["B1"] = "Montant"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Claire Martin"; ws["B2"] = 100
    ws["B3"] = "=B2*2"
    ws2 = wb.create_sheet("Tiers")
    ws2["A1"] = "Fournisseur Claire Martin"
    wb.save(path)

def test_masks_string_cells_all_sheets_preserves_formatting(tmp_path):
    src = tmp_path / "bal.xlsx"
    _make_book(src)
    ref = Referential.load_default()
    ner = FakeNer({"Claire Martin": "PERSON"})
    res = anonymize_xlsx(src, ner, ref, tmp_path,
                         when=datetime(2026, 6, 24, 17, 18, 0))
    assert res.output_path.name == "bal_ano_20260624171800.xlsx"
    wb = openpyxl.load_workbook(res.output_path)
    ws = wb["Balance"]
    assert ws["A2"].value == "[PERSONNE]"
    assert ws["B2"].value == 100
    assert ws["B3"].value == "=B2*2"
    assert ws["A1"].font.bold is True
    assert ws["A1"].value == "Libellé"
    assert wb["Tiers"]["A1"].value == "Fournisseur [PERSONNE]"
    assert openpyxl.load_workbook(src)["Balance"]["A2"].value == "Claire Martin"
    assert any(r["original"] == "Claire Martin" for r in res.report.to_rows())
